#!/usr/bin/python

## Program will create UPDATE, DELETE, or INSERT statements from one of:
##  - CSV File - unimplemented
##  - XLS File - unimplemented
##  - XLSX File - in progress

# CAVEATS:
#  - Program failed to parse file that I edited in LibreOffice. why?
#  - XLS file renamed to XLSX did not work with openpyxl

# Launch Web Browser w/ Data in File
#  - http://anh.cs.luc.edu/python/hands-on/3.1/handsonHtml/webtemplates.html


from Tkinter import *
from tkFileDialog import askopenfilename
from openpyxl.reader.excel import load_workbook
from openpyxl.cell import Cell
import sys

filename = ''

class MakeStatement(Tk):
    def __init__(self):
        Tk.__init__(self) # root widget, only 1 per program!
        self.wm_title("Make a Statement")
        frame = Frame(self)
        # frame.grid_propagate(0)
        frame.grid()
        self.loadFile = Button(self, text="Browse", command=self.getFileName)
        self.loadFile.grid(row=0,column=0)
        self.mainInfo = Label(self, text='for M$ Excel or CSV file.')
        self.mainInfo.grid(row=0,column=1)

    def getFileName(self):
        self.action = IntVar()
        self.filename = askopenfilename()
        self.mainInfo.grid_forget()
        self.loadFile.grid_forget()
        self.db = self.getData() # get
        numTables = len(self.db)
        numFields = []


        # list all the sheets (tables)
        self.tableInfo = Label(self, text='Select Table:')
        self.tableInfo.grid(row=2,column=0)
        self.tableList = Listbox(self)
        self.tableList.grid(row=3,column=0)
        
        self.whereInfo = Label(self, text='for M$ Excel or CSV file.')
        self.whereInfo.grid(row=0,column=1)
        
        for i in range(0,len(self.db)): # only for first sheet, need to think how to do recursively
            self.tableList.insert(END, self.db[i][0])

        # for i in range(0,len(self.db)):
        #     Label(self, text='Table: ' + self.db[i][0]).grid(row=2,column=i)

        # print str(len(self.db))
        # print str(self.db[1])

        self.fileLabel = Label(self, text="Loaded: .." + self.filename[-36:])
        self.fileLabel.grid(row=1,column=0,columnspan=2)

        Radiobutton(self, text="UPDATE", indicatoron=0, variable=self.action, value=1, command=self.whereClause).grid(row=4,column=0)
        Radiobutton(self, text="INSERT", indicatoron=0, variable=self.action, value=2, command=self.whereClause).grid(row=4,column=1)
        Radiobutton(self, text="DELETE", indicatoron=0, variable=self.action, value=3, command=self.whereClause).grid(row=4,column=2)
        go = Button(self, text="CONVERT", command=self.createHTML)
        go.grid(row=6,column=0,columnspan=3)


    # directs to parsing functions for each file type
    def getData(self):
        if self.filename[-5:] == '.xlsx':
            return self.loadXlsx()
        else:
            sys.exit()

    # creates fields/values tuple from an XLSX spreadsheet
    def loadXlsx(self):
        fields = [] # populates from the first row in the spreadsheet
        values = [] # all the values in sheet (not including first row)
        dataRow = [] # a single row of data in the spreadsheet
        table = [] # array of the tableName, then fields and finally values
        database = [] # the entire spreadsheet in multi-dimensional array; return value
        
        wb = load_workbook(self.filename)
        sheetNames = wb.get_sheet_names() # sheetNames is Array of sheet names
        sheets = []

        for i in range(0,len(sheetNames)):
            sheets.append(wb.get_sheet_by_name(sheetNames[i]))

        for i in range(0,len(sheets)):
            for rIndex in range(0,len(sheets[i].rows)):
                for cIndex in range(0,len(sheets[i].columns)):
                    if rIndex == 0: 
                        fields.append(str(sheets[i].cell(row=rIndex,column=cIndex).value))
                    else:
                        dataRow.append(str(sheets[i].cell(row=rIndex,column=cIndex).value))
                    if cIndex == (len(sheets[i].columns) - 1):
                        values.append(dataRow)
                        dataRow = []
                if rIndex == (len(sheets[i].rows) - 1):
                    while 'None' in fields: # clean up any BLANK fields
                        fields.remove('None')
                    table.append(sheetNames[i])
                    table.append(fields)
                    table.append(values)
                    fields = []
                    values = []

            database.append(table)
            table = []

        return database #return tuple of fields and values

    # Called when UPDATE is selected
    def whereClause(self):
        if self.action.get() == 1:
            self.whereInfo = Label(self, text='Update On:')
            self.whereInfo.grid(row=2,column=1)
            self.fieldList = Listbox(self, selectmode=MULTIPLE)
            self.fieldList.grid(row=3,column=1,columnspan=2)
            
            for field in self.db[0][1]: # only for first sheet, need to think how to do recursively
                self.fieldList.insert(END, field)
        else:
            if hasattr(self, 'fieldList'):
                self.fieldList.grid_forget()
                self.whereInfo.grid_forget()


    # Action Database -> String (of formatted html)
    def createHTML(self):
        print 'action: ' + str(self.action.get())



if __name__ == "__main__":
    root = MakeStatement()
    root.mainloop()


